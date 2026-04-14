import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

driver = webdriver.Chrome()
driver.implicitly_wait(5)
#downloading the excel from the website
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

driver.implicitly_wait(10)

#Editing the excel
book = openpyxl.load_workbook("C:\\Users\\Manasa\\collector\\Downloads\\download.xlsx")
sheet = book.active


for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        if sheet.cell(row=i, column=j).value == "Apple":
            sheet.cell(row=i, column=j+2).value = '900'

book.save("C:\\Users\\Manasa\\collector\\Downloads\\download.xlsx")


fruit_name = 'Apple'

driver.find_element(By.ID, "fileinput").send_keys("C:\\Users\\Manasa\\collector\\Downloads\\download.xlsx")
wait = WebDriverWait(driver, 5)
Locator = (By.CLASS_NAME, "Toastify__toast-body")
print(wait.until(expected_conditions.visibility_of_element_located(Locator)).text)

price = driver.find_element(By.XPATH, "//div[text() = 'Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text() = '"+fruit_name+"']/parent::div/parent::div/div[@id = 'cell-"+price+"-undefined']").text
print(actual_price)

