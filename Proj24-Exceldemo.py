import openpyxl

book = openpyxl.load_workbook("C:\\Excel\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row = 1, column=1)
print(cell.value)

sheet.cell(row=2, column=2).value = "Manasa"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['C1'].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row = i, column =1).value == "testcase2":
      for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)