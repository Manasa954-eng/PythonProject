import os.path
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

def update_excel_sheet(file_path, search_parameter, new_value):
 book = openpyxl.load_workbook(file_path)
 sheet = book.active

 for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        if sheet.cell(row=i, column=j).value == search_parameter:
            sheet.cell(row=i, column=j+2).value = new_value

 book.save("C:\\Users\\Manasa\\collector\\Downloads\\download.xlsx")


driver = webdriver.Chrome()
driver.implicitly_wait(5)

#downloading the excel from the website
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

while not os.path.exists("C:\\Users\\Manasa\\collector\\Downloads\\download.xlsx"):
    time.sleep(5)


#Editing the excel
fruit_name = 'Apple'
value = '116'
update_excel_sheet("C:\\Users\\Manasa\\collector\\Downloads\\download.xlsx",fruit_name, value)


driver.implicitly_wait(10)

#Uploading the excel
driver.find_element(By.ID, "fileinput").send_keys("C:\\Users\\Manasa\\collector\\Downloads\\download.xlsx")
wait = WebDriverWait(driver, 5)
Locator = (By.CLASS_NAME, "Toastify__toast-body")
print(wait.until(expected_conditions.visibility_of_element_located(Locator)).text)

price = driver.find_element(By.XPATH, "//div[text() = 'Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text() = '"+fruit_name+"']/parent::div/parent::div/div[@id = 'cell-"+price+"-undefined']").text
print(actual_price)
assert actual_price == value

